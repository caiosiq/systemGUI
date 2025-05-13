import time
import matplotlib.pyplot as plt
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import os
import threading

class Graph:
    def __init__(self, temperatures_dict, pressures_dict, balances_dict, flow_rates_dict, 
                 max_points=1000, update_interval=0.5):
        """
        Enhanced graph utility for real-time data visualization.
        
        Args:
            temperatures_dict: Dictionary of temperature data series
            pressures_dict: Dictionary of pressure data series
            balances_dict: Dictionary of balance data series  
            flow_rates_dict: Dictionary of flow rate data series
            max_points: Maximum number of data points to keep per series (default: 1000)
            update_interval: Time between plot updates in seconds (default: 0.5)
        """
        self.temperatures_dict = temperatures_dict
        self.pressures_dict = pressures_dict
        self.balances_dict = balances_dict
        self.flow_rates_dict = flow_rates_dict
        
        self.data_dicts = [
            ('Temperatures', self.temperatures_dict),
            ('Pressures', self.pressures_dict),
            ('Balances', self.balances_dict),
            ('Flow_Rates', self.flow_rates_dict)
        ]
        
        # Dictionary mapping plot types to their properties
        self.plot_properties = {
            'Temperatures': {'index': 0, 'ylabel': 'Temperature (°C)', 'color_map': 'inferno'},
            'Pressures': {'index': 1, 'ylabel': 'Pressure (psi)', 'color_map': 'viridis'},
            'Balances': {'index': 2, 'ylabel': 'Balance (g)', 'color_map': 'cividis'},
            'Flow_Rates': {'index': 3, 'ylabel': 'Flow_Rate (mL/min)', 'color_map': 'plasma'}
        }
        
        self.color_map = {}
        self.gui_plot_stopped = False
        self.max_points = max_points
        self.update_interval = update_interval
        self.last_update_time = time.time()
        
        # For tracking min/max values for each plot type
        self.value_ranges = {
            'Temperatures': {'min': float('inf'), 'max': float('-inf')},
            'Pressures': {'min': float('inf'), 'max': float('-inf')},
            'Balances': {'min': float('inf'), 'max': float('-inf')},
            'Flow_Rates': {'min': float('inf'), 'max': float('-inf')}
        }
        
        # For tracking time window
        self.start_time = None
        self.time_window = 120  # Default time window in seconds

    def toggle_all_series(self, dict_type):
        """
        Toggle visibility of all data series in a specific category.
        
        Args:
            dict_type: Type of dictionary to toggle ('temperatures', 'pressures', etc.)
        """
        d = self.get_dict_type(dict_type)
        if d:
            # Check if all are currently on or off
            all_on = all(d[name][0] for name in d)
            
            # Toggle all to the opposite state
            for name in d:
                d[name][0] = not all_on

    def plot(self, plots, canvas, fig=None):
        """
        Main plotting function - continuously updates plots with new data.
        
        Args:
            plots: List of matplotlib subplot axes
            canvas: Canvas to draw on
            fig: Figure object (optional)
        """
        if self.start_time is None:
            self.start_time = time.time()
            
        while not self.gui_plot_stopped:
            current_time = time.time()
            
            # Update plots at specified interval to reduce CPU usage
            if (current_time - self.last_update_time) < self.update_interval:
                time.sleep(0.05)  # Small sleep to prevent CPU hogging
                continue
                
            self.last_update_time = current_time
            
            # Clear plots but maintain settings
            for p in plots:
                p.clear()
            
            # Set up plots
            for label, properties in self.plot_properties.items():
                label = label[:-1] # remove s from the end of the label
                idx = properties['index']
                plots[idx].set_title(f'{label} Over Time')
                plots[idx].set_xlabel('Time (s)')
                plots[idx].set_ylabel(properties['ylabel'])
                plots[idx].grid(True, linestyle='--', alpha=0.7)
            
            # Plot data series
            for label, data_dict in self.data_dicts:
                plotted = False
                p_idx = self.plot_properties[label]['index']
                p = plots[p_idx]
                
                for name, var_value in data_dict.items():
                    if var_value[0] and var_value[1] and len(var_value[2]) > 0:
                        # Filter out None values and limit to max points
                        valid_data = [(t, v) for t, v in var_value[2][-self.max_points:] if t is not None and v is not None]
                        
                        if valid_data:
                            times = [(t - self.start_time) for t, v in valid_data]  # Relative time in seconds
                            values = [v for t, v in valid_data]
                            
                            # Update min/max values for auto-scaling
                            if values:
                                self.value_ranges[label]['min'] = min(self.value_ranges[label]['min'], min(values))
                                self.value_ranges[label]['max'] = max(self.value_ranges[label]['max'], max(values))
                            
                            # Assign consistent colors to each series
                            if name not in self.color_map:
                                cmap = plt.get_cmap(self.plot_properties[label]['color_map'])
                                self.color_map[name] = cmap(hash(name) % 10 / 10.0)
                            
                            line, = p.plot(times, values, label=f'{name}', color=self.color_map[name], 
                                          linewidth=2, marker='o', markersize=2, alpha=0.8)
                            
                            # Add annotation for the latest value if there are values
                            if values:
                                latest_idx = len(values) - 1
                                latest_value = values[latest_idx]
                                latest_time = times[latest_idx]
                                p.annotate(f'{latest_value:.2f}', 
                                          xy=(latest_time, latest_value),
                                          xytext=(4, 4), 
                                          textcoords='offset points',
                                          fontsize=8)
                            
                            plotted = True
                
                if plotted:
                    # Set better y limits with padding
                    if self.value_ranges[label]['min'] != float('inf'):
                        data_range = self.value_ranges[label]['max'] - self.value_ranges[label]['min']
                        padding = max(0.1 * data_range, 0.1)  # At least 0.1 unit padding
                        p.set_ylim(
                            self.value_ranges[label]['min'] - padding,
                            self.value_ranges[label]['max'] + padding
                        )
                    
                    p.legend(loc='upper right', fontsize=8)
            
            # Adjust the time window - dynamic based on available data
            latest_times = []
            for label, data_dict in self.data_dicts:
                for name, var_value in data_dict.items():
                    if var_value[0] and var_value[1] and var_value[2]:
                        valid_times = [t for t, v in var_value[2] if t is not None]
                        if valid_times:
                            latest_times.append(max(valid_times))
            
            if latest_times:
                latest_time = max(latest_times)
                time_range = latest_time - self.start_time
                
                # Update all x-axes to show the same time range
                for p in plots:
                    # Show at most the last time_window seconds of data
                    if time_range > self.time_window:
                        p.set_xlim(time_range - self.time_window, time_range + 2)
                    else:
                        p.set_xlim(0, max(time_range + 2, self.time_window))
            
            # Make plots look good
            if fig:
                fig.tight_layout()
            
            # Draw the canvas
            canvas.draw()
            
            # Small sleep to prevent CPU hogging
            time.sleep(0.05)

    def update_dict(self, dict_type, name, value):
        """
        Add a new data point to the specified dictionary.
        
        Args:
            dict_type: Type of dictionary to update
            name: Name of the data series
            value: New data value
        """
        if value != None:
            d = self.get_dict_type(dict_type)
            if d and name in d and d[name][0] and d[name][1]:
                # If we've reached max points, use a deque-like behavior
                if len(d[name][2]) >= self.max_points:
                    d[name][2] = d[name][2][-(self.max_points-1):] + [(time.time(), value)]
                else:
                    d[name][2].append((time.time(), value))

    def toggle_series(self, dict_type, name, is_visible=None):
        """
        Set the visibility of a specific data series.
        
        Args:
            dict_type: Type of dictionary containing the series
            name: Name of the data series to toggle
            is_visible: If provided, set to this visibility state; if None, toggle current state
        """
        d = self.get_dict_type(dict_type)
        if d and name in d:
            # If is_visible is provided, set to that state
            if is_visible is not None:
                # If turning off, add a discontinuity marker
                if d[name][1] and not is_visible:
                    d[name][2].append((None, None))
                d[name][1] = is_visible
            # Otherwise toggle current state
            else:
                # If turning off, add a discontinuity marker
                if d[name][1]:
                    d[name][2].append((None, None))
                d[name][1] = not d[name][1]
                
    def set_all_series(self, dict_type, is_visible):
        """
        Set visibility for all series in a dictionary.
        
        Args:
            dict_type: Type of dictionary to update
            is_visible: Boolean visibility state to set
        """
        d = self.get_dict_type(dict_type)
        if d:
            for name in d:
                # If turning off and currently visible, add discontinuity marker
                if d[name][1] and not is_visible:
                    d[name][2].append((None, None))
                d[name][1] = is_visible

    def get_dict_type(self, dict_type):
        """
        Get the dictionary corresponding to the specified type.
        
        Args:
            dict_type: Type of dictionary to get ('temperatures', 'pressures', etc.)
            
        Returns:
            Dictionary object or None if not found
        """
        return getattr(self, f"{dict_type.lower()}_dict", None)

    def set_time_window(self, seconds):
        """
        Set the time window to display (in seconds).
        
        Args:
            seconds: Number of seconds to display
        """
        self.time_window = seconds

    def export_data(self, filename=None):
        """
        Export all current data to a nicely formatted Excel file without timestamp column.
        Only includes data sheets for each measurement type.
        
        Args:
            filename: Output filename (default: "system2_data_YYYY-MM-DD_HH-MM-SS.xlsx")
            
        Returns:
            The filename of the exported file
        """
        if filename is None:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"system2_data_{timestamp}.xlsx"
        
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        
        # Create separate sheets for each data type
        data_types = ["Temperatures", "Pressures", "Balances", "Flow_Rates"]
        
        # Dictionary to hold timestamps for each data type
        all_timestamps_by_type = {}
        
        # First, collect all timestamps for each data type
        for data_type in data_types:
            dict_name = f"{data_type.lower()}_dict"
            data_dict = getattr(self, dict_name)
            
            # Collect all timestamps for this data type
            timestamps = set()
            for name, var_value in data_dict.items():
                if var_value[0] and var_value[1]:  # If series is active and visible
                    timestamps.update([t for t, v in var_value[2] if t is not None])
            
            all_timestamps_by_type[data_type] = sorted(timestamps)
        
        # Now create and populate each sheet
        for i, data_type in enumerate(data_types):
            # Create a sheet for this data type
            if i == 0:  # Use the default sheet for the first data type
                ws = wb.active
                ws.title = data_type
            else:
                ws = wb.create_sheet(title=data_type)
            
            # Reference the dictionary
            dict_name = f"{data_type.lower()}_dict"
            data_dict = getattr(self, dict_name)
            
            # Get sorted timestamps
            timestamps = all_timestamps_by_type[data_type]
            
            # Style configuration
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # First row: Headers - now we'll only have Date/Time column and data columns
            # Add date/time column header
            ws.cell(row=1, column=1).value = "Date/Time"
            ws.cell(row=1, column=1).font = header_font
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=1, column=1).alignment = header_alignment
            ws.cell(row=1, column=1).border = thin_border
            
            # Add data series headers
            col_idx = 2
            series_indices = {}  # To track column indices for each series
            
            for name in data_dict:
                if data_dict[name][0]:  # If the series is active
                    ws.cell(row=1, column=col_idx).value = name
                    ws.cell(row=1, column=col_idx).font = header_font
                    ws.cell(row=1, column=col_idx).fill = header_fill
                    ws.cell(row=1, column=col_idx).alignment = header_alignment
                    ws.cell(row=1, column=col_idx).border = thin_border
                    series_indices[name] = col_idx
                    col_idx += 1
            
            # Data rows
            for row_idx, ts in enumerate(timestamps, start=2):
                # Convert timestamp to readable datetime
                dt = datetime.datetime.fromtimestamp(ts)
                formatted_dt = dt.strftime("%Y-%m-%d %H:%M:%S")
                
                # Add datetime - no raw timestamp anymore
                ws.cell(row=row_idx, column=1).value = formatted_dt
                ws.cell(row=row_idx, column=1).border = thin_border
                
                # Add values for each series
                for name, col_idx in series_indices.items():
                    # Find the value at this timestamp
                    value = None
                    for t, v in data_dict[name][2]:
                        if t == ts:
                            value = v
                            break
                    
                    # Add to the sheet
                    if value is not None:
                        ws.cell(row=row_idx, column=col_idx).value = value
                        ws.cell(row=row_idx, column=col_idx).number_format = '0.0000'  # Format as number with 4 decimals
                    ws.cell(row=row_idx, column=col_idx).border = thin_border
            
            # Auto-adjust column widths
            for col_idx in range(1, len(series_indices) + 2):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            
            # Freeze the header row
            ws.freeze_panes = "A2"
        
        # Save the workbook
        try:
            wb.save(filename)
            return filename
        except PermissionError:
            # If file is open in another program, create a new filename
            base_name, ext = os.path.splitext(filename)
            new_filename = f"{base_name}_new{ext}"
            wb.save(new_filename)
            return new_filename

    def clear_data(self, dict_type=None, name=None):
        """
        Clear data points for a specific series or all series.
        
        Args:
            dict_type: Type of dictionary to clear (if None, clear all)
            name: Name of the series to clear (if None, clear all in dict_type)
        """
        if dict_type is None:
            # Clear all data
            for label, data_dict in self.data_dicts:
                for name in data_dict:
                    data_dict[name][2] = []
            # Reset value ranges
            for plot_type in self.value_ranges:
                self.value_ranges[plot_type] = {'min': float('inf'), 'max': float('-inf')}
            
        elif name is None:
            # Clear all data in the specified dictionary
            d = self.get_dict_type(dict_type)
            if d:
                for name in d:
                    d[name][2] = []
                # Reset value range for this type
                plot_type = dict_type.capitalize()
                self.value_ranges[plot_type] = {'min': float('inf'), 'max': float('-inf')}
        else:
            # Clear only the specified series
            d = self.get_dict_type(dict_type)
            if d and name in d:
                d[name][2] = []
                # We'll recalculate min/max on next update

    def stop_plotting(self, stop=True):
        """
        Stop or resume the plotting loop.
        
        Args:
            stop: True to stop plotting, False to resume
        """
        self.gui_plot_stopped = stop

class DataCollector:
    """
    A class to synchronize data collection and ensure all sensors are read together
    before recording data points. This addresses the issue of partially empty rows.
    """
    def __init__(self, graph):
        """
        Initialize the DataCollector with reference to the graph object.
        
        Args:
            graph: The Graph object where data is stored
        """
        self.graph = graph
        self.collection_interval = 1.0  # 1 second interval
        self.running = False
        self.thread = None
        
        # Buffers to store most recent sensor values
        self.temperature_buffer = {}
        self.pressure_buffer = {}
        self.flow_rate_buffer = {}
        self.balance_buffer = {}
        
        # Lock to protect buffers during updates
        self.buffer_lock = threading.Lock()
    
    def buffer_update(self, data_type, name, value):
        """
        Updates the appropriate buffer with the latest sensor reading.
        This method should be called by the sensor reading callbacks.
        
        Args:
            data_type: Type of data (temperatures, pressures, etc.)
            name: Name of the sensor
            value: Current sensor reading
        """
        with self.buffer_lock:
            if data_type == "temperatures":
                self.temperature_buffer[name] = value
            elif data_type == "pressures":
                self.pressure_buffer[name] = value
            elif data_type == "flow_rates":
                self.flow_rate_buffer[name] = value
            elif data_type == "balances":
                self.balance_buffer[name] = value
    
    def start_collection(self):
        """Start the synchronized data collection thread."""
        if not self.running:
            self.running = True
            self.thread = threading.Thread(target=self._collection_loop)
            self.thread.daemon = True
            self.thread.start()
            return True
        return False
    
    def stop_collection(self):
        """Stop the data collection thread."""
        self.running = False
        if self.thread and self.thread.is_alive():
            self.thread.join(timeout=2.0)
    
    def _collection_loop(self):
        """
        Main collection loop that runs at a fixed interval.
        This ensures all data points are collected at the same timestamp.
        """
        while self.running:
            # Get the current timestamp
            timestamp = time.time()
            
            # Capture current buffer state to avoid race conditions
            with self.buffer_lock:
                temp_data = self.temperature_buffer.copy()
                press_data = self.pressure_buffer.copy()
                flow_data = self.flow_rate_buffer.copy()
                bal_data = self.balance_buffer.copy()
            
            # Update the graph with synchronized data
            for name, value in temp_data.items():
                if self.graph.temperatures_dict.get(name) and self.graph.temperatures_dict[name][0] and self.graph.temperatures_dict[name][1]:
                    # Replace the automatic timestamp with our synchronized one
                    self.graph.temperatures_dict[name][2].append((timestamp, value))
            
            for name, value in press_data.items():
                if self.graph.pressures_dict.get(name) and self.graph.pressures_dict[name][0] and self.graph.pressures_dict[name][1]:
                    self.graph.pressures_dict[name][2].append((timestamp, value))
            
            for name, value in flow_data.items():
                if self.graph.flow_rates_dict.get(name) and self.graph.flow_rates_dict[name][0] and self.graph.flow_rates_dict[name][1]:
                    self.graph.flow_rates_dict[name][2].append((timestamp, value))
            
            for name, value in bal_data.items():
                if self.graph.balances_dict.get(name) and self.graph.balances_dict[name][0] and self.graph.balances_dict[name][1]:
                    self.graph.balances_dict[name][2].append((timestamp, value))
            
            # Sleep for the collection interval
            time.sleep(self.collection_interval)